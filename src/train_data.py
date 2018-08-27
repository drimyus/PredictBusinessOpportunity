import sys
import os
import csv
import datetime
from openpyxl import load_workbook


TARGET_CATEGORY = 'Ease of Doing Business Ranking'
TITLES = ['Geography', 'Category', 'Data', 'Type', 'Unit', 'Current', 'Constant', '2012', '2013', '2014', '2015',
          '2016', '2017', '2018', '2019', '2020']

_cur_dir = os.path.dirname(os.path.realpath(__file__))


class TrainData:
    def __init__(self, debug=False):
        now = datetime.datetime.now()
        self.cur_year = now.year
        self.debug = debug

    @staticmethod
    def read_xlsx(xlsx_path, sheet_idx=0):
        workbook = load_workbook(xlsx_path, data_only=True)
        sheet_name = workbook.get_sheet_names()[sheet_idx]
        worksheet = workbook.get_sheet_by_name(sheet_name)

        raw_data = []
        for row in worksheet.iter_rows():
            line = []
            for cell in row:
                line.append(cell.value)
            raw_data.append(line)

        return raw_data, sheet_name

    def cleaning(self, raw_data):
        # check the first line titles ---------------------------------------------------------------------
        value_cols_idxs = []
        category_col_idx = -1
        first_raw = raw_data[0]
        for title in first_raw:
            try:
                if title.isdigit() and self.cur_year > int(title) > 2000:
                    value_cols_idxs.append(first_raw.index(title))

                if title.lower() == 'Category'.lower():
                    category_col_idx = (first_raw.index(title))
            except Exception as e:
                print(e)
                break

        target_row_idx = -1
        for i in range(1, len(raw_data)):
            line = raw_data[i]
            if line[category_col_idx].replace(' ', '').lower() == TARGET_CATEGORY.replace(' ', '').lower():
                target_row_idx = i
                break

        if self.debug:
            sys.stdout.write("category_col_idx: {}\n".format(category_col_idx))
            sys.stdout.write("target_row_idx: {}\n".format(target_row_idx))

        # crop the value area -----------------------------------------------------------------
        x_data = []
        y_data = []
        for i in range(1, len(raw_data)):
            try:
                if i == target_row_idx:
                    for j in value_cols_idxs:
                        y_data.append(float(raw_data[i][j]))
                    continue

                value_line = []
                for j in value_cols_idxs:
                    value_line.append(float(raw_data[i][j]))

                x_data.append(value_line)
            except Exception as e:
                print(e)

        # swap rows and cols on x_data ----------------------------------------------------------------------------
        swap_x_data = []
        for j in range(len(x_data[0])):
            new_line = []
            for i in range(len(x_data)):
                new_line.append(x_data[i][j])
            swap_x_data.append(new_line)

        years = [raw_data[0][j] for j in value_cols_idxs]
        categories = [raw_data[i][category_col_idx] for i in range(1, len(raw_data))]

        return swap_x_data, y_data, years, categories

    @staticmethod
    def normalizing(x_data, y_data):
        max_vales = [-10000] * len(x_data[0])
        for line in x_data:
            for j in range(len(line)):
                if max_vales[j] < abs(line[j]):
                    max_vales[j] = abs(line[j])

        norm_data = []
        for i in range(len(x_data)):
            line = x_data[i]
            new_line = []
            for j in range(len(line)):
                new_line.append(line[j] / max_vales[j])

            new_line.append(y_data[i])

            norm_data.append(new_line)

        return norm_data, max_vales

    @staticmethod
    def write_train_data(norm_data, years, categories, max_values):
        # write the train_data.csv file on the same location -------------------------------------------
        train_data_dir = os.path.join(_cur_dir, os.pardir, "data/train")

        train_data_path = os.path.join(train_data_dir, "train_data.csv")
        if sys.version_info[0] == 2:  # py 2x
            with open(train_data_path, 'wb') as fp:  # for python 2x
                wr = csv.writer(fp, delimiter=',')
                wr.writerows(norm_data)
        elif sys.version_info[0] == 3:  # py 3x
            with open(train_data_path, 'w', newline='') as fp:  # for python 3x
                wr = csv.writer(fp, delimiter=',')
                wr.writerows(norm_data)

        # write the years.txt on the same location ------------------------------------------------------
        year_label_path = os.path.join(train_data_dir, "years.txt")
        with open(year_label_path, 'w') as fp:
            for year in years:
                fp.write(year + "\n")

        # write the categories.txt on the same location ------------------------------------------------------
        categories_label_path = os.path.join(train_data_dir, "categories.txt")
        with open(categories_label_path, 'w') as fp:
            for category in categories:
                fp.write(category + "\n")

        sys.stdout.write("Create the train_data.csv successfully!\n")
        return train_data_path

    def train_data(self, path):

        raw_data, sheet_name = self.read_xlsx(xlsx_path=path)
        if self.debug:
            sys.stdout.write("sheet name: {}\n".format(sheet_name))

        x_data, y_data, years, categories = self.cleaning(raw_data=raw_data)
        if self.debug:
            sys.stdout.write("years: {}\n".format(years))
            sys.stdout.write("categories: {}\n".format(categories))

            sys.stdout.write("x_data: {}\n")
            for line in x_data:
                print(line)

            sys.stdout.write("y_data: {}\n")
            for line in y_data:
                print(line)

        norm_data, max_values = self.normalizing(x_data=x_data, y_data=y_data)

        self.write_train_data(norm_data=norm_data, years=years, categories=categories, max_values=max_values)


if __name__ == '__main__':
    _path = "../data/raw_data/boi_analysis(1).xlsx"
    td = TrainData()
    td.train_data(path=_path)
